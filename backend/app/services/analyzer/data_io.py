"""生信数据多格式导入导出"""
import csv
import io
import json
from typing import Any, Dict


class BioDataIO:
    """生信数据导入导出工具"""

    @staticmethod
    async def import_csv(content: bytes, has_header: bool = True) -> Dict[str, Any]:
        """CSV 导入 — 基因表达矩阵
        格式：第一列为基因名，后续列为样本表达值
        """
        text = content.decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return {"genes": {}, "samples": []}
        header = rows[0] if has_header else [f"sample_{i}" for i in range(len(rows[0]) - 1)]
        samples = header[1:]
        genes = {}
        for row in rows[1:] if has_header else rows:
            if len(row) < 2:
                continue
            gene = row[0]
            values = [float(v) if v else 0.0 for v in row[1:]]
            genes[gene] = values
        return {"genes": genes, "samples": samples, "format": "csv"}

    @staticmethod
    async def import_json(content: bytes) -> Dict[str, Any]:
        """JSON 导入 — {genes: {gene: [values]}, samples: [...]}"""
        data = json.loads(content.decode("utf-8"))
        return {"genes": data.get("genes", {}), "samples": data.get("samples", []), "format": "json"}

    @staticmethod
    async def import_fasta(content: bytes) -> Dict[str, Any]:
        """FASTA 导入 — 序列数据"""
        text = content.decode("utf-8")
        sequences = []
        current_id = None
        current_seq = []
        for line in text.splitlines():
            line = line.strip()
            if line.startswith(">"):
                if current_id is not None:
                    joined = "".join(current_seq)
                    sequences.append({"id": current_id, "seq": joined, "length": len(joined)})
                current_id = line[1:].split()[0] if len(line) > 1 else f"seq_{len(sequences)}"
                current_seq = []
            elif line:
                current_seq.append(line)
        if current_id is not None:
            joined = "".join(current_seq)
            sequences.append({"id": current_id, "seq": joined, "length": len(joined)})
        return {"sequences": sequences, "count": len(sequences), "format": "fasta"}

    @staticmethod
    async def import_vcf(content: bytes) -> Dict[str, Any]:
        """VCF 导入 — 变异信息（简化版）"""
        text = content.decode("utf-8")
        variants = []
        for line in text.splitlines():
            if line.startswith("#") or not line.strip():
                continue
            fields = line.split("\t")
            if len(fields) < 5:
                continue
            variants.append({
                "chrom": fields[0], "pos": int(fields[1]) if fields[1].isdigit() else 0,
                "id": fields[2], "ref": fields[3], "alt": fields[4],
            })
        return {"variants": variants, "count": len(variants), "format": "vcf"}

    @staticmethod
    async def import_mtx(content: bytes) -> Dict[str, Any]:
        """MTX 导入 — 稀疏矩阵（Market Matrix 格式简化版）"""
        text = content.decode("utf-8")
        lines = [l for l in text.splitlines() if l.strip() and not l.startswith("%")]
        if not lines:
            return {"matrix": [], "format": "mtx"}
        header = lines[0].split()
        rows, cols, nnz = int(header[0]), int(header[1]), int(header[2])
        matrix = []
        for line in lines[1:1 + nnz]:
            parts = line.split()
            if len(parts) >= 3:
                matrix.append({"row": int(parts[0]), "col": int(parts[1]), "value": float(parts[2])})
        return {"matrix": matrix, "shape": [rows, cols], "nnz": nnz, "format": "mtx"}

    @staticmethod
    async def import_excel(content: bytes) -> Dict[str, Any]:
        """Excel 多 sheet 导入 — 使用 openpyxl 读取

        Args:
            content: Excel 文件二进制内容（.xlsx/.xls）
        Returns:
            {sheets: {sheet_name: {headers, rows}}, format: "excel"}
            openpyxl 未安装时返回 {error: "..."}
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {"error": "openpyxl 未安装，请运行 pip install openpyxl"}

        wb = load_workbook(io.BytesIO(content), read_only=True)
        sheets: Dict[str, Any] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
            data_rows = []
            for row in rows[1:]:
                if row is None or all(c is None for c in row):
                    continue
                data_rows.append(dict(zip(header, row)))
            sheets[sheet_name] = {"headers": header, "rows": data_rows, "count": len(data_rows)}
        wb.close()
        return {"sheets": sheets, "format": "excel", "sheet_count": len(sheets)}

    @staticmethod
    async def export_excel(data: Dict[str, Any]) -> bytes:
        """Excel 导出 — 分析结果写入单 sheet

        Args:
            data: 分析结果字典（支持 genes/clusters/pathways 结构）
        Returns:
            Excel 文件二进制内容
        Raises:
            ImportError: openpyxl 未安装
        """
        try:
            from openpyxl import Workbook
        except ImportError as e:
            raise ImportError("openpyxl 未安装，请运行 pip install openpyxl") from e

        wb = Workbook()
        ws = wb.active
        ws.title = "analysis_result"

        if "genes" in data:
            ws.append(["gene", "log2fc", "pvalue", "padj", "regulation", "significant"])
            for g in data["genes"]:
                ws.append([g.get("gene"), g.get("log2fc"), g.get("pvalue"),
                           g.get("padj"), g.get("regulation"), g.get("significant")])
        elif "clusters" in data:
            ws.append(["gene", "cluster_id", "pca_x", "pca_y"])
            for c in data["clusters"]:
                ws.append([c.get("gene"), c.get("cluster_id"), c.get("pca_x"), c.get("pca_y")])
        elif "pathways" in data:
            ws.append(["id", "name", "pvalue", "ratio", "genes"])
            for p in data["pathways"]:
                ws.append([p.get("id"), p.get("name"), p.get("pvalue"),
                           p.get("ratio"), ";".join(p.get("genes", []))])
        elif "samples" in data:
            ws.append(["sample", "pc1", "pc2"])
            for s in data["samples"]:
                ws.append([s.get("sample"), s.get("pc1"), s.get("pc2")])
        else:
            ws.append(["key", "value"])
            for k, v in data.items():
                ws.append([k, str(v)[:200]])

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    async def export_csv(data: Dict[str, Any]) -> bytes:
        """CSV 导出 — 差异表达/聚类/通路结果"""
        output = io.StringIO()
        writer = csv.writer(output)
        if "genes" in data:
            writer.writerow(["gene", "log2fc", "pvalue", "padj", "regulation", "significant"])
            for g in data["genes"]:
                writer.writerow([g.get("gene"), g.get("log2fc"), g.get("pvalue"),
                                 g.get("padj"), g.get("regulation"), g.get("significant")])
        elif "clusters" in data:
            writer.writerow(["gene", "cluster_id", "pca_x", "pca_y"])
            for c in data["clusters"]:
                writer.writerow([c.get("gene"), c.get("cluster_id"), c.get("pca_x"), c.get("pca_y")])
        elif "pathways" in data:
            writer.writerow(["id", "name", "pvalue", "ratio", "genes"])
            for p in data["pathways"]:
                writer.writerow([p.get("id"), p.get("name"), p.get("pvalue"),
                                 p.get("ratio"), ";".join(p.get("genes", []))])
        else:
            writer.writerow(["key", "value"])
            for k, v in data.items():
                writer.writerow([k, str(v)[:100]])
        return output.getvalue().encode("utf-8")

    @staticmethod
    async def export_json(data: Dict[str, Any]) -> bytes:
        """JSON 导出 — 任意分析结果"""
        return json.dumps(data, ensure_ascii=False, indent=2, default=str).encode("utf-8")

    @staticmethod
    async def detect_format(filename: str, content: bytes) -> str:
        """自动检测文件格式"""
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext in ("csv", "tsv", "json", "fasta", "fa", "vcf", "mtx", "xlsx", "xls"):
            return "tsv" if ext == "tsv" else ext
        head = content[:200].decode("utf-8", errors="ignore").strip()
        if head.startswith(">"):
            return "fasta"
        if head.startswith("##fileformat=VCF"):
            return "vcf"
        if head.startswith("{"):
            return "json"
        return "csv"


__all__ = ["BioDataIO"]
